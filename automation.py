import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# Loading the Excel workbook and saving it as 'wb'.
wb = xl.load_workbook('auto.xlsx')

# Accessing the sheet with the data of interest.
sheet = wb['Sheet1']

# Accessing the first cell on the sheet.
cell = sheet['A1']

# Looping through the sheet and calculating the expected result.
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

# Creating a reference to be put into a bar chart.
values = Reference(sheet,
                   min_row=2,
                   max_row=sheet.max_row,
                   min_col=4,
                   max_col=4)

# Creating a bar chart from the reference
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e1')

# Saving the workbook object into an Excel file called 'transaction2.xlsx'.
wb.save('transaction2.xlsx')
